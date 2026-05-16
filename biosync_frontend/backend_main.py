from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from openpyxl import load_workbook
import json
import os
import subprocess
from datetime import datetime, date, timedelta
from collections import defaultdict

app = FastAPI()

# ================= CORS =================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ================= BASE PATH =================

BASE_DIR = r"C:\biosync_web_app"
SCRIPT_DIR = os.path.join(BASE_DIR, "scripts")
CONFIG_DIR = os.path.join(BASE_DIR, "config")

REPORT_CONFIG = os.path.join(CONFIG_DIR, "report_config.json")
SDK_CONFIG = os.path.join(CONFIG_DIR, "sdk_config.json")
SDK_LAST_STATUS = os.path.join(CONFIG_DIR, "sdk_last_status.json")
DEPARTMENT_USERS_FILE = os.path.join(BASE_DIR, "department_users.json")

ATTENDANCE_MASTER_FILE = r"C:\AttendanceAutomation\attendance_master.xlsx"
BACKUP_FOLDER = r"C:\AttendanceAutomation\BackupExcel"

# ================= MODELS =================

class LoginData(BaseModel):
    username: str
    password: str

class DepartmentRangeReportRequest(BaseModel):
    department: str
    start_date: str
    end_date: str

class ParentAttendanceRequest(BaseModel):
    reg_no: str
    year: str
    class_name: str
    start_date: str
    end_date: str

# ================= DEPARTMENT SHEET MAP =================

DEPARTMENT_SHEETS = {
    "CSE": [
        "I CSE A", "I CSE B", "I CSE C",
        "II CSE A", "II CSE B", "II CSE C",
        "III CSE A", "III CSE B", "III CSE C",
        "IV CSE A", "IV CSE B", "IV CSE C"
    ],
    "CSBS": [
        "I CSBS", "II CSBS", "III CSBS", "IV CSBS"
    ],
    "AIML": [
        "I AIML A", "I AIML B", "II AIML A", "II AIML B", "III AIML"
    ],
    "IT": [
        "I IT A", "I IT B", "I IT C",
        "II IT A", "II IT B", "II IT C",
        "III IT A", "III IT B", "III IT C",
        "IV IT A", "IV IT B", "IV IT C"
    ],
    "MECH": [
        "I MECH", "II MECH", "III MECH", "IV MECH"
    ],
    "CIVIL": [
        "I CIVIL", "II CIVIL", "III CIVIL", "IV CIVIL"
    ],
    "CHEMICAL": [
        "I CHEMICAL", "II CHEMICAL", "III CHEMICAL", "IV CHEMICAL"
    ],
    "EEE": [
        "I EEE", "II EEE", "III EEE", "IV EEE"
    ],
    "CCE": [
        "I CCE", "II CCE", "III CCE", "IV CCE"
    ],
    "BIOTECH": [
        "I BIOTECH", "II BIOTECH", "III BIOTECH", "IV BIOTECH"
    ],
    "BME": [
        "I BME", "II BME", "III BME", "IV BME"
    ],
    "AIDS": [
        "I AIDS A", "I AIDS B", "I AIDS C", "I AIDS D",
        "II AIDS A", "II AIDS B", "II AIDS C", "II AIDS D",
        "III AIDS A", "III AIDS B",
        "IV AIDS A", "IV AIDS B"
    ],
    "ECE": [
        "I ECE A", "I ECE B", "I ECE C", "I ECE D", "I ECE E", "I ECE F",
        "II ECE A", "II ECE B", "II ECE C", "II ECE D",
        "III ECE A", "III ECE B", "III ECE C", "III ECE D",
        "IV ECE A", "IV ECE B", "IV ECE C"
    ]
}

# ================= HELPERS =================

def safe_percentage(present: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round((present / total) * 100, 2)

def is_present_value(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() == "present"

def get_no_cache_headers():
    return {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0"
    }

def read_json_file(file_path, default_value):
    if not os.path.exists(file_path):
        return default_value

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_value

def build_sdk_status_payload():
    default_status = {
        "successful_devices": [],
        "offline_devices": [],
        "no_log_devices": [],
        "error_devices": [],
        "last_run": None
    }

    if not os.path.exists(SDK_LAST_STATUS):
        return default_status

    try:
        with open(SDK_LAST_STATUS, "r", encoding="utf-8") as f:
            data = json.load(f)

        return {
            "successful_devices": data.get("successful_devices", []),
            "offline_devices": data.get("offline_devices", []),
            "no_log_devices": data.get("no_log_devices", []),
            "error_devices": data.get("error_devices", []),
            "last_run": data.get("last_run")
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read SDK last status: {str(e)}")

def get_existing_department_sheet_map(workbook):
    existing_sheetnames = set(workbook.sheetnames)
    filtered_map = {}

    for dept, sheets in DEPARTMENT_SHEETS.items():
        existing = [sheet for sheet in sheets if sheet in existing_sheetnames]
        filtered_map[dept] = existing

    return filtered_map

def count_sheet_attendance(ws):
    present = 0
    absent = 0
    total = 0

    for row in range(2, ws.max_row + 1):
        student_value = ws.cell(row, 2).value
        status_value = ws.cell(row, 10).value

        if student_value is None or str(student_value).strip() == "":
            continue

        total += 1

        if is_present_value(status_value):
            present += 1
        else:
            absent += 1

    return {
        "present": present,
        "absent": absent,
        "total": total
    }

def sum_department_attendance(workbook, department_name: str):
    dept_key = department_name.strip().upper()
    department_map = get_existing_department_sheet_map(workbook)

    if dept_key not in department_map:
        raise HTTPException(status_code=404, detail=f"Department '{department_name}' not found")

    present = 0
    absent = 0
    total = 0

    for sheet_name in department_map[dept_key]:
        ws = workbook[sheet_name]
        sheet_data = count_sheet_attendance(ws)
        present += sheet_data["present"]
        absent += sheet_data["absent"]
        total += sheet_data["total"]

    return {
        "name": dept_key,
        "present": present,
        "absent": absent,
        "total": total,
        "attendance_percentage": safe_percentage(present, total)
    }

def sum_college_attendance(workbook):
    department_map = get_existing_department_sheet_map(workbook)

    present = 0
    absent = 0
    total = 0
    counted_sheets = set()

    for _, sheets in department_map.items():
        for sheet_name in sheets:
            if sheet_name in counted_sheets:
                continue

            counted_sheets.add(sheet_name)
            ws = workbook[sheet_name]
            sheet_data = count_sheet_attendance(ws)
            present += sheet_data["present"]
            absent += sheet_data["absent"]
            total += sheet_data["total"]

    return {
        "present": present,
        "absent": absent,
        "total": total
    }

def build_department_comparison(workbook, requested_department: str):
    department_map = get_existing_department_sheet_map(workbook)
    comparison = []

    for dept_name in department_map.keys():
        dept_data = sum_department_attendance(workbook, dept_name)
        comparison.append(dept_data)

    requested_key = requested_department.strip().upper()

    comparison.sort(
        key=lambda item: (
            0 if item["name"] == requested_key else 1,
            item["name"]
        )
    )

    return comparison

def parse_input_date(date_str: str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Date format must be YYYY-MM-DD")

def find_backup_files_in_range(start_date, end_date):
    if not os.path.exists(BACKUP_FOLDER):
        raise HTTPException(status_code=500, detail="BackupExcel folder missing")

    files = []

    for root, dirs, filenames in os.walk(BACKUP_FOLDER):
        for filename in filenames:
            if not filename.lower().endswith(".xlsx"):
                continue

            try:
                file_date = datetime.strptime(filename.lower().replace(".xlsx", ""), "%d-%m-%Y").date()
            except Exception:
                continue

            if file_date.weekday() == 6:
                continue

            if start_date <= file_date <= end_date:
                files.append((file_date, os.path.join(root, filename)))

    files.sort(key=lambda x: x[0])
    return files

def get_existing_sheets_for_department_in_workbook(workbook, department_name: str):
    dept_key = department_name.strip().upper()
    if dept_key not in DEPARTMENT_SHEETS:
        raise HTTPException(status_code=404, detail=f"Department '{department_name}' not found")

    return [sheet for sheet in DEPARTMENT_SHEETS[dept_key] if sheet in workbook.sheetnames]

def count_present_total_for_sheet(ws):
    present = 0
    total = 0

    for row in range(2, ws.max_row + 1):
        student_value = ws.cell(row, 2).value
        status_value = ws.cell(row, 10).value

        if student_value is None or str(student_value).strip() == "":
            continue

        total += 1
        if is_present_value(status_value):
            present += 1

    return present, total

def normalize_student_key(reg_no, student_name):
    reg = "" if reg_no is None else str(reg_no).strip()
    name = "" if student_name is None else str(student_name).strip()
    return reg if reg else name

def extract_class_students_for_single_day(ws):
    students = []

    for row in range(2, ws.max_row + 1):
        reg_no = ws.cell(row, 2).value
        if reg_no is None or str(reg_no).strip() == "":
            continue

        student_name = ws.cell(row, 4).value
        status_value = ws.cell(row, 10).value

        present_days = 1 if is_present_value(status_value) else 0
        total_days = 1
        absent_days = total_days - present_days

        students.append({
            "student_name": "" if student_name is None else str(student_name).strip(),
            "reg_no": "" if reg_no is None else str(reg_no).strip(),
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "attendance_percentage": safe_percentage(present_days, total_days)
        })

    students.sort(key=lambda item: (item["student_name"], item["reg_no"]))
    return students

def build_department_range_report_for_today(dept_key: str, report_date: date):
    if not os.path.exists(ATTENDANCE_MASTER_FILE):
        raise HTTPException(status_code=500, detail="attendance_master.xlsx missing")

    try:
        wb = load_workbook(ATTENDANCE_MASTER_FILE, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open attendance master file: {str(e)}")

    try:
        existing_class_sheets = get_existing_sheets_for_department_in_workbook(wb, dept_key)

        class_comparison = []
        class_breakdown = []
        student_breakdown = {}
        day_present_total = 0
        day_total_total = 0

        for class_name in existing_class_sheets:
            ws = wb[class_name]
            present, total = count_present_total_for_sheet(ws)
            attendance_percentage = safe_percentage(present, total)

            class_comparison.append({
                "class_name": class_name,
                "attendance_percentage": attendance_percentage
            })

            class_breakdown.append({
                "class_name": class_name,
                "total_students": total,
                "present_average": round(present, 2),
                "attendance_percentage": attendance_percentage
            })

            student_breakdown[class_name] = extract_class_students_for_single_day(ws)

            day_present_total += present
            day_total_total += total

        overall_average = safe_percentage(day_present_total, day_total_total)

        daily_trend = []
        if day_total_total > 0:
            daily_trend.append({
                "date": report_date.strftime("%Y-%m-%d"),
                "attendance_percentage": overall_average
            })

        class_comparison.sort(key=lambda x: x["class_name"])
        class_breakdown.sort(key=lambda x: x["class_name"])

        return {
            "department": dept_key,
            "start_date": report_date.strftime("%Y-%m-%d"),
            "end_date": report_date.strftime("%Y-%m-%d"),
            "overall_average": overall_average,
            "class_comparison": class_comparison,
            "daily_trend": daily_trend,
            "class_breakdown": class_breakdown,
            "student_breakdown": student_breakdown
        }

    finally:
        wb.close()

def build_department_range_report_from_backup(dept_key: str, start_date, end_date):
    files = find_backup_files_in_range(start_date, end_date)

    class_stats = {
        class_name: {
            "present_sum": 0,
            "total_sum": 0,
            "days_found": 0,
            "student_counts": []
        }
        for class_name in DEPARTMENT_SHEETS[dept_key]
    }

    student_stats = {
        class_name: {}
        for class_name in DEPARTMENT_SHEETS[dept_key]
    }

    daily_trend = []

    for file_date, file_path in files:
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception:
            continue

        existing_class_sheets = get_existing_sheets_for_department_in_workbook(wb, dept_key)

        day_present = 0
        day_total = 0

        for class_name in existing_class_sheets:
            ws = wb[class_name]
            present, total = count_present_total_for_sheet(ws)

            class_stats[class_name]["present_sum"] += present
            class_stats[class_name]["total_sum"] += total
            class_stats[class_name]["days_found"] += 1
            class_stats[class_name]["student_counts"].append(total)

            day_present += present
            day_total += total

            for row in range(2, ws.max_row + 1):
                reg_no = ws.cell(row, 2).value
                if reg_no is None or str(reg_no).strip() == "":
                    continue

                student_name = ws.cell(row, 4).value
                status_value = ws.cell(row, 10).value

                student_key = normalize_student_key(reg_no, student_name)

                if student_key not in student_stats[class_name]:
                    student_stats[class_name][student_key] = {
                        "student_name": "" if student_name is None else str(student_name).strip(),
                        "reg_no": "" if reg_no is None else str(reg_no).strip(),
                        "total_days": 0,
                        "present_days": 0
                    }

                student_stats[class_name][student_key]["total_days"] += 1
                if is_present_value(status_value):
                    student_stats[class_name][student_key]["present_days"] += 1

        if day_total > 0:
            daily_trend.append({
                "date": file_date.strftime("%Y-%m-%d"),
                "attendance_percentage": safe_percentage(day_present, day_total)
            })

        wb.close()

    filtered_class_stats = {
        class_name: stats
        for class_name, stats in class_stats.items()
        if stats["total_sum"] > 0
    }

    total_present_all = sum(stats["present_sum"] for stats in filtered_class_stats.values())
    total_possible_all = sum(stats["total_sum"] for stats in filtered_class_stats.values())
    overall_average = safe_percentage(total_present_all, total_possible_all)

    class_comparison = []
    class_breakdown = []
    student_breakdown = {}

    for class_name, stats in filtered_class_stats.items():
        attendance_percentage = safe_percentage(stats["present_sum"], stats["total_sum"])

        if stats["student_counts"]:
            total_students = max(stats["student_counts"])
        else:
            total_students = 0

        present_average = round(
            stats["present_sum"] / stats["days_found"], 2
        ) if stats["days_found"] > 0 else 0.0

        class_comparison.append({
            "class_name": class_name,
            "attendance_percentage": attendance_percentage
        })

        class_breakdown.append({
            "class_name": class_name,
            "total_students": total_students,
            "present_average": present_average,
            "attendance_percentage": attendance_percentage
        })

        class_students = []
        for student in student_stats.get(class_name, {}).values():
            total_days = student["total_days"]
            present_days = student["present_days"]
            absent_days = total_days - present_days

            class_students.append({
                "student_name": student["student_name"],
                "reg_no": student["reg_no"],
                "total_days": total_days,
                "present_days": present_days,
                "absent_days": absent_days,
                "attendance_percentage": safe_percentage(present_days, total_days)
            })

        class_students.sort(key=lambda item: (item["student_name"], item["reg_no"]))
        student_breakdown[class_name] = class_students

    class_comparison.sort(key=lambda x: x["class_name"])
    class_breakdown.sort(key=lambda x: x["class_name"])
    daily_trend.sort(key=lambda x: x["date"])

    return {
        "department": dept_key,
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date": end_date.strftime("%Y-%m-%d"),
        "overall_average": overall_average,
        "class_comparison": class_comparison,
        "daily_trend": daily_trend,
        "class_breakdown": class_breakdown,
        "student_breakdown": student_breakdown
    }

def build_department_range_report(department: str, start_date_str: str, end_date_str: str):
    dept_key = department.strip().upper()
    start_date = parse_input_date(start_date_str)
    end_date = parse_input_date(end_date_str)

    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date cannot be greater than end_date")

    if dept_key not in DEPARTMENT_SHEETS:
        raise HTTPException(status_code=404, detail=f"Department '{department}' not found")

    today = date.today()

    if start_date == end_date == today:
        return build_department_range_report_for_today(dept_key, today)

    return build_department_range_report_from_backup(dept_key, start_date, end_date)

# ================= ADMIN LOGIN =================

@app.post("/admin/login")
def admin_login(data: LoginData):

    file_path = os.path.join(BASE_DIR, "admin_users.json")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=500, detail="admin_users.json missing")

    with open(file_path, "r", encoding="utf-8") as f:
        admins = json.load(f)

    for admin in admins:
        if admin["username"] == data.username and admin["password"] == data.password:
            return {"status": "success"}

    raise HTTPException(status_code=401, detail="Invalid login")

# ================= DEPARTMENT LOGIN =================

@app.post("/department/login")
def department_login(data: LoginData):

    if not os.path.exists(DEPARTMENT_USERS_FILE):
        raise HTTPException(status_code=500, detail="department_users.json missing")

    try:
        with open(DEPARTMENT_USERS_FILE, "r", encoding="utf-8") as f:
            users = json.load(f)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read department_users.json: {str(e)}")

    for user in users:
        if user.get("username") == data.username and user.get("password") == data.password:
            return {
                "status": "success",
                "role": "department",
                "department": user.get("department"),
                "display_name": user.get("display_name", user.get("department"))
            }

    raise HTTPException(status_code=401, detail="Invalid department login")

# ================= DEPARTMENT DASHBOARD SUMMARY =================

@app.get("/department/dashboard-summary/{department}")
def get_department_dashboard_summary(department: str):

    if not os.path.exists(ATTENDANCE_MASTER_FILE):
        raise HTTPException(status_code=500, detail="attendance_master.xlsx missing")

    try:
        wb = load_workbook(ATTENDANCE_MASTER_FILE, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open attendance master file: {str(e)}")

    try:
        college_data = sum_college_attendance(wb)
        department_data = sum_department_attendance(wb, department)
        comparison_data = build_department_comparison(wb, department)

        return {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "college": college_data,
            "department": department_data,
            "comparison": comparison_data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build department dashboard summary: {str(e)}")
    finally:
        wb.close()

# ================= DEPARTMENT RANGE REPORT =================

@app.post("/department/range-report")
def department_range_report(data: DepartmentRangeReportRequest):
    try:
        return build_department_range_report(
            department=data.department,
            start_date_str=data.start_date,
            end_date_str=data.end_date
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build department range report: {str(e)}")

# ================= SDK CONFIG =================

@app.get("/sdk/config")
def get_sdk_config():
    headers = get_no_cache_headers()
    config_data = read_json_file(SDK_CONFIG, {})
    return JSONResponse(content=config_data, headers=headers)

@app.put("/sdk/config")
def update_sdk_config(data: dict):

    os.makedirs(CONFIG_DIR, exist_ok=True)

    with open(SDK_CONFIG, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

    return {"status": "updated"}

# ================= SDK LAST STATUS =================

@app.get("/sdk/last-status")
def get_sdk_last_status():
    headers = get_no_cache_headers()
    status_payload = build_sdk_status_payload()
    return JSONResponse(content=status_payload, headers=headers)

# ================= SDK BOOTSTRAP =================

@app.get("/sdk/bootstrap")
def get_sdk_bootstrap():
    headers = get_no_cache_headers()

    config_data = read_json_file(SDK_CONFIG, {})
    status_payload = build_sdk_status_payload()

    return JSONResponse(
        content={
            "config": config_data,
            "status": status_payload
        },
        headers=headers
    )

# ================= RUN SDK =================

@app.post("/run/sdk_pull")
def run_sdk():

    script = os.path.join(SCRIPT_DIR, "sdk_pull_attendance.py")

    if not os.path.exists(script):
        raise HTTPException(status_code=500, detail="SDK script missing")

    subprocess.Popen(
        f'start cmd /k python "{script}"',
        shell=True
    )

    return {"status": "sdk started"}

# ================= RUN COMBINE =================

@app.post("/run/combine")
def run_combine():

    script = os.path.join(SCRIPT_DIR, "combine_all_sheets.py")

    if not os.path.exists(script):
        raise HTTPException(status_code=500, detail="Combine script missing")

    subprocess.Popen(
        f'start cmd /k python "{script}"',
        shell=True
    )

    return {"status": "combine started"}

# ================= REPORT CONFIG =================

@app.get("/report/config")
def get_report_config():

    if not os.path.exists(REPORT_CONFIG):
        return {
            "department": None,
            "class_name": None,
            "emp_code": None,
            "start_date": None,
            "end_date": None
        }

    with open(REPORT_CONFIG, "r", encoding="utf-8") as f:
        return json.load(f)

@app.put("/report/config")
def update_report_config(data: dict):

    os.makedirs(CONFIG_DIR, exist_ok=True)

    config = {
        "department": data.get("department") or None,
        "class_name": data.get("class_name") or None,
        "emp_code": data.get("emp_code") or None,
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date")
    }

    with open(REPORT_CONFIG, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4)

    return {
        "status": "config_saved",
        "config": config
    }

# ================= RUN RANGE REPORT =================

@app.post("/run/report")
def run_report():

    script = os.path.join(SCRIPT_DIR, "biosync_range_report.py")

    if not os.path.exists(script):
        raise HTTPException(status_code=500, detail="Report engine missing")

    subprocess.Popen(
        f'start cmd /k python "{script}"',
        shell=True
    )

    return {"status": "report started"}

# ================= PARENT DASHBOARD =================

def normalize_reg_no(val):
    if val is None:
        return ""
    val_str = str(val).strip().replace(" ", "")
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str.strip()

def get_possible_reg_keys(input_reg_no):
    normalized = normalize_reg_no(input_reg_no)
    keys = {normalized}
    if len(normalized) > 8:
        keys.add(normalized[-8:])
    if normalized.startswith("9225"):
        keys.add(normalized[4:])
    return list(keys)

def search_student_in_target_sheet(wb, possible_keys, search_reg_no, target_date_str, target_sheet, department):
    target_sheet_normalized = target_sheet.strip().upper()
    
    actual_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().upper() == target_sheet_normalized:
            actual_sheet_name = name
            break
            
    if actual_sheet_name:
        print(f"[DEBUG] Target sheet '{actual_sheet_name}' exists in {target_date_str}")
        ws = wb[actual_sheet_name]
        
        sample_regs = []
        for row in range(2, min(ws.max_row + 1, 7)):
            reg = ws.cell(row, 2).value
            if reg is not None:
                sample_regs.append(normalize_reg_no(reg))
        print(f"[DEBUG] First few normalized regs in column B: {sample_regs}")
        
        for row in range(2, ws.max_row + 1):
            reg_no = ws.cell(row, 2).value
            if reg_no is not None:
                norm_excel_reg = normalize_reg_no(reg_no)
                if norm_excel_reg in possible_keys:
                    print(f"[DEBUG] Match found! row {row}, norm_excel: {norm_excel_reg}")
                    student_name = ws.cell(row, 4).value
                    punch_time = ws.cell(row, 5).value
                    status_value = ws.cell(row, 10).value
                    print(f"[DEBUG] Matched student_name: {student_name}, status: {status_value}, punch_time: {punch_time}")
                    return {
                        "date": target_date_str,
                        "student_name": "" if student_name is None else str(student_name).strip(),
                        "reg_no": search_reg_no,
                        "class_name": actual_sheet_name,
                        "department": department,
                        "punch_time": "-" if punch_time is None else str(punch_time).strip(),
                        "status": "Present" if is_present_value(status_value) else "Absent"
                    }
        print(f"[DEBUG] Register number NOT matched in target sheet '{actual_sheet_name}'")
    else:
        print(f"[DEBUG] Target sheet '{target_sheet_normalized}' does NOT exist in {target_date_str}")
        print(f"[DEBUG] Available sheets: {wb.sheetnames}")
    return None

@app.post("/parent/student-attendance")
def parent_student_attendance(data: ParentAttendanceRequest):
    search_reg_no = normalize_reg_no(data.reg_no)
    possible_keys = get_possible_reg_keys(data.reg_no)
    target_sheet = f"{data.year.strip()} {data.class_name.strip()}"
    department = data.class_name.strip().split()[0] if data.class_name.strip() else ""
    
    print(f"[DEBUG] Received payload: {data.dict()}")
    print(f"[DEBUG] Normalized reg_no: {search_reg_no}")
    print(f"[DEBUG] Possible register keys: {possible_keys}")
    print(f"[DEBUG] Target sheet name: {target_sheet}")
    try:
        try:
            start_date = datetime.strptime(data.start_date, "%Y-%m-%d").date()
        except ValueError:
            start_date = datetime.strptime(data.start_date, "%d-%m-%Y").date()
            
        try:
            end_date = datetime.strptime(data.end_date, "%Y-%m-%d").date()
        except ValueError:
            end_date = datetime.strptime(data.end_date, "%d-%m-%Y").date()
            
        print(f"[DEBUG] Parsed start_date: {start_date}, end_date: {end_date}")
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD or DD-MM-YYYY")
        
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date cannot be greater than end_date")
        
    today = date.today()
    
    daily_records = []
    found_student_info = None
    target_sheet_found_anywhere = False

    backup_files = find_backup_files_in_range(start_date, end_date)
    backup_file_dict = {fd: fp for fd, fp in backup_files}

    current_date = start_date
    dates_to_check = []
    while current_date <= end_date:
        if current_date.weekday() != 6:
            dates_to_check.append(current_date)
        current_date += timedelta(days=1)

    for dt in dates_to_check:
        print(f"[DEBUG] Checking date: {dt}")
        file_path = None
        if dt == today:
            file_path = ATTENDANCE_MASTER_FILE
        else:
            file_path = backup_file_dict.get(dt)
            
        if not file_path:
            print(f"[DEBUG] File missing for date {dt}")
            continue
            
        print(f"[DEBUG] File path found: {file_path}")
        if not os.path.exists(file_path):
            print(f"[DEBUG] File does not exist on disk: {file_path}")
            continue
            
        try:
            wb = load_workbook(file_path, data_only=True)
            target_sheet_normalized = target_sheet.strip().upper()
            if any(name.strip().upper() == target_sheet_normalized for name in wb.sheetnames):
                target_sheet_found_anywhere = True
            record = search_student_in_target_sheet(wb, possible_keys, data.reg_no, dt.strftime("%Y-%m-%d"), target_sheet, department)
            if record:
                if not found_student_info:
                    found_student_info = record
                daily_records.append(record)
            wb.close()
        except Exception as e:
            print(f"[DEBUG] Failed to process {file_path}: {str(e)}")
            continue

    if not target_sheet_found_anywhere:
        raise HTTPException(status_code=404, detail="No class sheet found for the selected year and class in the selected range.")

    if not found_student_info:
        raise HTTPException(status_code=404, detail="No attendance record found for this register number, year, class, and selected range.")

    total_days = len(daily_records)
    present_days = sum(1 for r in daily_records if r["status"] == "Present")
    absent_days = total_days - present_days
    pct = safe_percentage(present_days, total_days)

    if pct > 90:
        status_label = "Excellent"
    elif pct >= 80:
        status_label = "Good"
    elif pct >= 75:
        status_label = "Needs Attention"
    else:
        status_label = "Bad / Critical"

    chart_trend = []
    monthly_trend_dict = defaultdict(list)
    for r in daily_records:
        val = 100 if r["status"] == "Present" else 0
        chart_trend.append({
            "date": r["date"],
            "value": val
        })
        try:
            r_date = parse_input_date(r["date"])
            month_label = r_date.strftime("%B %Y")
            monthly_trend_dict[month_label].append({
                "date": r["date"],
                "value": val
            })
        except Exception:
            pass

    return {
        "reg_no": search_reg_no,
        "student_name": found_student_info["student_name"],
        "class_name": found_student_info["class_name"],
        "department": found_student_info["department"],
        "start_date": data.start_date,
        "end_date": data.end_date,
        "summary": {
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "attendance_percentage": pct,
            "status_label": status_label
        },
        "daily_records": daily_records,
        "chart_data": {
            "present_absent": [
                { "name": "Present", "value": present_days },
                { "name": "Absent", "value": absent_days }
            ],
            "daily_trend": chart_trend
        },
        "monthly_trend": dict(monthly_trend_dict)
    }